import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side

from output_config import RESULT_DIR

class XlDecorator():
    def __init__(self,result_file_name,jissui_month_list):
        
        
        
        self.file_path = str(RESULT_DIR)+result_file_name
        
        self.file_path = f"{str(RESULT_DIR)}/{result_file_name}"
        
        self.file_path = RESULT_DIR / result_file_name
        
        
        
        
        #sheet_name_list = ["超過月末在庫月数(平均販売量に対して)"]
        self.jissui_month_list = jissui_month_list
        
    
    def decorate_breakdown_time(self):
        wb = load_workbook(self.file_path)

        # シート名でシートを選択
        sheet = wb['工程時間内訳']  #シートオブジェクト
        
        non_productive_time_fill = PatternFill(start_color='EBF1DE',
                                               end_color='EBF1DE',
                                               fill_type='solid')
        non_productive_string_list = ["保全","開発テスト","生技テスト","計画停止","切替時間","立下","立上"]
        
        meta_time_fill = PatternFill(start_color='DA9694',
                                    end_color='DA9694',
                                    fill_type='solid')
        
        meta_string_list = ["暦日時間","負荷時間","余力時間"]
        
        
        productive_time_fill = PatternFill(start_color='FDE9D9',
                                    end_color='FDE9D9',
                                    fill_type='solid')
        
        
        
        
        # シートの全行と全列を走査
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value in non_productive_string_list:
                    cell.fill = non_productive_time_fill
                if cell.value in meta_string_list:
                    cell.fill = meta_time_fill
                if (cell.value not in non_productive_string_list) and (cell.value not in meta_string_list):
                    cell.fill = productive_time_fill
                
                if cell.value == "余力時間":
                    #余力時間セルから横に罫線引っ張る
                    row = cell.row  # セルの行番号
                    column = cell.column  # セルの列番号
                     # 太い下罫線を定義
                    border = Border(bottom=Side(border_style="thick"))  # 下罫線を太くする
                    
                    # 見つかったセルから15列目まで罫線を引く
                    for col_offset in range(0, len(self.jissui_month_list)+2):  # 0から14の範囲で15列目まで
                        sheet.cell(row=row, column=column + col_offset).border = border
                
                
                if (cell.value == "保全") or (cell.value == "負荷時間"):
                    #余力時間セルから横に罫線引っ張る
                    row = cell.row  # セルの行番号
                    column = cell.column  # セルの列番号
                     # 太い下罫線を定義
                    border = Border(top=Side(border_style="thin"))  # 下罫線を太くする
                    
                    # 見つかったセルから15列目まで罫線を引く
                    for col_offset in range(0, len(self.jissui_month_list)+2):  # 0から14の範囲で15列目まで
                        sheet.cell(row=row, column=column + col_offset).border = border
                
        
        
        
                    
        # 変更を上書き保存
        wb.save(self.file_path)
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    def main(self):
        
        self.decorate_breakdown_time()
        
        
        
        
        # ワークブックを読み込む
        wb = load_workbook(self.file_path)

        # シート名でシートを選択
        sheet = wb['超過月末在庫月数(平均販売量に対して)']  #シートオブジェクト
        
        
        # セルを走査して、値が正の場合は背景色を黄色に設定
        #黄色：FFFF00
        #オレンジ：FFC000
        yellow_fill = PatternFill(start_color='FFC000',
                                end_color='FFC000',
                                fill_type='solid')

        # シートの全行と全列を走査
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                # セルの値が数値であり、かつ正の場合のみ背景を黄色にする
                if isinstance(cell.value, (int, float)) and cell.value > 0:
                    cell.fill = yellow_fill

        # 変更を上書き保存
        wb.save(self.file_path)


################################################################################################

if __name__ == "__main__":
    file_path = "./結果/結果_20240404_132426.xlsx"
    XlDecorator(file_path).main()
        
        