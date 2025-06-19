import datetime
from openpyxl import load_workbook,Workbook
import pandas as pd
#import streamlit as st
import pickle
import numpy as np
from openpyxl.styles import Font, Border, Side


from output_config import RESULT_DIR


class ResultOutputter():
    """
    最適化結果出力クラス
    
    
    
    """
    
    def __init__(self,all_params_dict,variables, problem,
                 filename):
        
        
        self.jissui_month_list = all_params_dict["jissui_month_list"]
        #self.fuka_dict = all_params_dict["fuka_dict"]
        #self.ave_switch_dict = all_params_dict["ave_switch_dict"]                
        self.sales_dict = all_params_dict["sales_dict"]
        self.width_dict = all_params_dict["width_dict"]
        self.cs_dict = all_params_dict["cs_dict"]
        self.finalized_sales_dict = all_params_dict["finalized_sales_dict"]
        self.finalized_prod_dict = all_params_dict["finalized_prod_dict"]
        self.init_stock_dict = all_params_dict["init_stock_dict"]
        self.min_continuous_dict = all_params_dict["min_continuous_dict"]
        self.prod_num_times_dict = all_params_dict["prod_num_times_dict"]
        self.basic_stock_min_dict = all_params_dict["basic_stock_min_dict"]
        self.basic_stock_max_dict = all_params_dict["basic_stock_max_dict"]     #基準在庫Max
        self.achieve_rate_dict = all_params_dict["achieve_rate_dict"]
        #self.all_index = self.get_all_index()
        self.all_index = all_params_dict["all_index"]

        self.plant_prod_dict = all_params_dict["plant_prod_dict"]
        #self.plant_prod_dict = self.get_plant_prod_dict()
        self.plant_list = list(self.plant_prod_dict.keys())
        self.prod_plant_dict = all_params_dict["prod_plant_dict"]
        #self.prod_plant_dict = self.get_prod_plant_dict()
        
        self.std_fc_dict = all_params_dict["std_fc_dict"]
        self.std_vc_dict = all_params_dict["std_vc_dict"]
        
        
        
        self.all_prod_list = list(self.cs_dict.keys()) 

        
        self.prod_amount_dict = variables.x                                           #生産量（千平米）
        self.sales_amount_dict = variables.y                                         #販売量（千平米）
        self.switch_time_dict = variables.switch_time                                           #各工場各月内切り替え時間
        
        #TODO月間切替
        self.inter_switch_time_head_dict = variables.inter_switch_time_head
        self.inter_switch_time_tail_dict = variables.inter_switch_time_tail
        
        
        self.filename = filename
        self.ave_sales_info_dict = all_params_dict["ave_sales_info_dict"]#
        self.ave_month_num = self.ave_sales_info_dict["ave_month_num"]       #在庫月数計算時に販売量何か月平均するか
        self.ave_sales_mode = self.ave_sales_info_dict["ave_sales_mode"]     #平均するときに当月を含むかどうか
        self.ave_month_dict = self.get_ave_month_dict(self.ave_month_num,mode=self.ave_sales_mode)
        self.next_month_dict = self.get_ave_month_dict(1)
        
        self.sales_prod_order = all_params_dict["sales_prod_order"]                            #合計販売量のシートの品種カラムと同じ順序のリスト
        self.cs_prod_order = all_params_dict["cs_prod_order"]                                  #csのシートの品種カラムと同じ順序のリスト
        
        
        self.prod_capacity_dict = all_params_dict["prod_capacity_dict"]                        #生産能力
        
        
        
        self.bundle_maint_dict = all_params_dict["bundle_maint_dict"]                          #保全抱き合わせの工場
        
        
        #self.cs_custom_order_ranking = {value: rank for rank, value in enumerate(self.cs_prod_order)}        #CSの品種順
        self.sales_custom_order_ranking = {value: rank for rank, value in enumerate(self.sales_prod_order)}  #キー品種名、value出現する順番
        
        #self.result_dir = "./結果/"
        
        self.result_file_path = RESULT_DIR / filename
        
        
        
        self.meta_params_info_dict = {"パラメータファイル名称":all_params_dict["params_file_name"], 
                                      "使用した制約条件":all_params_dict["constraint_list"], 
                                      "制約条件の適用期間":all_params_dict["constraint_period_dict"],
                                      "制約条件の適用工場":all_params_dict["constraint_plant_dict"],
                                      "使用した目的関数":all_params_dict["obj_priority_dict"]}
        
        
        
        
        self.main_dope_prod_dict = all_params_dict["main_dope_prod_dict"]   #メインドープ・メイン品種
        
        
        

    def get_filename(self):
        """
        結果ファイル名取得
        
        """
        now = datetime.datetime.now()
        filename = "結果_"+now.strftime('%Y%m%d_%H%M%S')+".xlsx"
        
        return filename
    
    
    
    def get_ave_month_dict(self,window_num,mode="含まない"):
        """
        販売量の平均をとる月のリストを辞書で。
        {4月:[5月,6月,7月],
        5月:[6月,7月,8月]}
        
        """
        
        ave_month_dict = {}
        
        if mode == "含む":
            diff = 0
        if mode == "含まない":
            diff = 1
        
        
        for i in range(len(self.jissui_month_list)):
            
            if i == len(self.jissui_month_list) - 1:
                month_list = [self.jissui_month_list[i]]    #最終月は最終月自身で。
            
            else:
                try:
                    month_list = self.jissui_month_list[i+diff:i+diff+window_num]
                except:
                    month_list = self.jissui_month_list[i+diff:]
            
            ave_month_dict[self.jissui_month_list[i]] = month_list
        
        return ave_month_dict
    
    
    
    def get_sorted_df(self,df,order_mode="cs"):
        """
        
        csや合計販売量のパラメタファイルと同じ順になるようにソートされたdfにする
        
        
        """
        if order_mode == "cs":
            #df['カスタムランク'] = df['品種名'].map(self.cs_custom_order_ranking)
            df['カスタムランク'] = df.apply(lambda row: self.cs_prod_order.get((row['品種名'], row['工場'])), axis=1)
            #df = df.sort_values(by=["工場", 'カスタムランク'])
            df = df.sort_values(by=['カスタムランク'])
            sorted_df = df.drop('カスタムランク', axis=1)
            sorted_df = sorted_df.reset_index(drop=True)
            return sorted_df
        
        
        
        
        if order_mode == "sales":
            df['カスタムランク'] = df['品種名'].map(self.sales_custom_order_ranking)
            df = df.sort_values(by='カスタムランク')
            sorted_df = df.drop('カスタムランク', axis=1)
            sorted_df = sorted_df.reset_index(drop=True)
            return sorted_df
            
        
            
        
    
    def convert_dict_to_df(self,amount_dict):
        """
        生産量や販売量の辞書データをデータフレームに変換
        
        """
        record_dict = {}
        df_list = []
        for prod_name in self.all_prod_list:
            record_dict["品種名"] = prod_name
            for plant in self.prod_plant_dict[prod_name]:
                record_dict["工場"] = plant
                for month in self.jissui_month_list:
                    record_dict[month] = [amount_dict[plant,month,prod_name].value()]
                    
                df_list.append(pd.DataFrame(record_dict))

        df_all = pd.concat(df_list)
        df_all[self.jissui_month_list] = df_all[self.jissui_month_list].apply(pd.to_numeric, errors='coerce')
        
        
        df_all = self.get_sorted_df(df_all)
        
        
        return df_all
    
    
    
    def convert_dict_to_df_params(self,params):
        """
        結果の出力に必要なパラメータ類のでーたフレーム化
        
        cs 基準在庫
        """
        record_dict = {}
        df_list = []
        for prod_name in self.all_prod_list:
            record_dict["品種名"] = prod_name
            for plant in self.prod_plant_dict[prod_name]:
                record_dict["工場"] = plant
                for month in self.jissui_month_list:
                    record_dict[month] = [params[prod_name][plant][month]]
                    
                df_list.append(pd.DataFrame(record_dict))

        df_all = pd.concat(df_list)
        df_all[self.jissui_month_list] = df_all[self.jissui_month_list].apply(pd.to_numeric, errors='coerce')  #
        
        df_all = self.get_sorted_df(df_all)
        
        return df_all
    
    def convert_dict_to_df_stock(self,params):
        """
        期首在庫
        
        """
        record_dict = {}
        df_list = []
        for prod_name in self.all_prod_list:
            record_dict["品種名"] = prod_name
            for plant in self.prod_plant_dict[prod_name]:
                record_dict["工場"] = plant
                record_dict["期首在庫"] = [params[prod_name][plant]]
                    
                df_list.append(pd.DataFrame(record_dict))

        df_all = pd.concat(df_list)
        df_all["期首在庫"] = df_all["期首在庫"].apply(pd.to_numeric, errors='coerce')
        
        df_all = self.get_sorted_df(df_all)
        
        return df_all
        


    def get_cumsum_df(self,df):
        """
        累積和のdfにする
        
        """
        df_cumsum = df.copy()
        df_cumsum[self.jissui_month_list] = df[self.jissui_month_list].cumsum(axis=1)
        return df_cumsum
        
    
    
    def get_stock_amount_df(self,prod_amount_df,sales_amount_df,init_stock_df):
        """
        在庫（千平米）

        Args:
            prod_amount_df (_type_): _description_
            sales_amount_df (_type_): _description_
            init_stock_df (_type_): _description_

        Returns:
            _type_: _description_
        """
        
        
        prod_amount_df_cumsum = self.get_cumsum_df(prod_amount_df)     #累積生産量
        sales_amount_df_cumsum = self.get_cumsum_df(sales_amount_df)   #累積販売量
        

        stock_amount_df = sales_amount_df.copy()
        
        stock_amount_df[self.jissui_month_list] = prod_amount_df_cumsum[self.jissui_month_list] \
                                                - sales_amount_df_cumsum[self.jissui_month_list] \
                                                    
        for month in self.jissui_month_list:
            stock_amount_df[month] = stock_amount_df[month] + init_stock_df['期首在庫']
        
        return stock_amount_df
    
    
    def convert_daynum_to_monthnum(self,basic_stock_df):
        """
        日数を月数に変換する関数。そもそも月数になったらいらんけど。
        
        ⇒ ver1.6.0以降いらなくなりました（笑）

        Args:
            basic_stock_df (_type_): _description_

        Returns:
            _type_: _description_
        """
        basic_stock_monthnum_df = basic_stock_df.copy()
        for month in self.jissui_month_list:
            basic_stock_monthnum_df[month] = basic_stock_df[month] /30
        
        return basic_stock_monthnum_df
    
    
    
    
        
        
    def convert_amount_to_num(self,df):
        """
        千平米から本数に換算する関数（3900ｍ換算）
        
        """
        
        num_df = df.copy() 
        for index in range(len(df)):
            prod_name = df["品種名"][index]
            num_df.loc[index,self.jissui_month_list] = df.loc[index,self.jissui_month_list] *1000/self.width_dict[prod_name]/3900
        
        return num_df
        
    def convert_amount_to_hour(self,df,cs_df,achieve_rate_df):
        """
        千平米から時間に換算
        
        """
        #achieve_rate_df = pd.DataFrame.from_dict(self.achieve_rate_dict)
        hour_df = df.copy()

        for index in range(len(df)):
            prod_name = df["品種名"][index]
            plant_name = df["工場"][index]
            
            
            hour_df.loc[index,self.jissui_month_list] = df.loc[index,self.jissui_month_list] *1000/\
            self.width_dict[prod_name]/cs_df.loc[index,self.jissui_month_list]/60/achieve_rate_df.loc[index,self.jissui_month_list]
        
        return hour_df
    

    
    
    
    
    def output_prod_amount_df(self,prod_amount_df):
        prod_amount_df = self.get_sorted_df(prod_amount_df)
        prod_amount_df=prod_amount_df.round(0)
        self.output_excel(prod_amount_df,sheet_name="生産量")
    

        
    
    
    def get_params_df(self):
        """
        パラメータの中で必要なものをdfに変換して取得
        
        """
        cs_df = self.convert_dict_to_df_params(self.cs_dict)                                 #CS
        basic_stock_min_df = self.convert_dict_to_df_params(self.basic_stock_min_dict)       #基準在庫月数Min
        basic_stock_max_df = self.convert_dict_to_df_params(self.basic_stock_max_dict)       #基準在庫月数Max
        init_stock_df = self.convert_dict_to_df_stock(self.init_stock_dict)                  #期首在庫
        std_fc_df = self.convert_dict_to_df_params(self.std_fc_dict)                         #標準FC単価
        std_vc_df = self.convert_dict_to_df_params(self.std_vc_dict)                         #標準VC単価
        achieve_rate_df = self.convert_dict_to_df_params(self.achieve_rate_dict)             #生産量達成率
        #st.write("生産量達成率だよ")
        #st.write(achieve_rate_df)
        
        
        return cs_df, basic_stock_min_df, basic_stock_max_df, init_stock_df, std_fc_df,std_vc_df,achieve_rate_df
    
    
    def get_amount_df(self,init_stock_df):
        """
        千平米で生産量、販売量、在庫を取得
        10桁レベル、7桁レベル。
        
        #TODO
        品種の順序と数字の桁数の確認
        生産量達成率が考慮されているか確認
        
        
        """
        prod_amount_df = self.convert_dict_to_df(self.prod_amount_dict)                   #生産量のdf
        sales_amount_df = self.convert_dict_to_df(self.sales_amount_dict)                 #販売量のdf
        stock_amount_df = self.get_stock_amount_df(prod_amount_df,
                                                   sales_amount_df,
                                                   init_stock_df)                         #在庫のdf
        
        prod_amount_df_7digits = prod_amount_df.groupby("品種名",as_index=False).sum()
        prod_amount_df_7digits = self.get_sorted_df(prod_amount_df_7digits,order_mode="sales")
        
        sales_amount_df_7digits = sales_amount_df.groupby("品種名",as_index=False).sum()
        sales_amount_df_7digits = self.get_sorted_df(sales_amount_df_7digits,order_mode="sales")
        
        stock_amount_df_7digits = stock_amount_df.groupby("品種名",as_index=False).sum()
        stock_amount_df_7digits = self.get_sorted_df(stock_amount_df_7digits,order_mode="sales")
        
        
        
        return prod_amount_df, sales_amount_df, stock_amount_df,\
               prod_amount_df_7digits, sales_amount_df_7digits, stock_amount_df_7digits
    
    
    
    def get_num_df(self,prod_amount_df,sales_amount_df,stock_amount_df):
        """
        本数のdfを取得       
        

        Args:
            prod_amount_df (_type_): _description_
            sales_amount_df (_type_): _description_
            stock_amount_df (_type_): _description_

        Returns:
            _type_: _description_
        """
        prod_num_df = self.convert_amount_to_num(prod_amount_df)              #生産本数のdf
        sales_num_df = self.convert_amount_to_num(sales_amount_df)            #販売本数のdf
        stock_num_df = self.convert_amount_to_num(stock_amount_df)            #在庫本数のdf
        
        prod_num_df_7digits = prod_num_df.groupby("品種名",as_index=False).sum()
        prod_num_df_7digits = self.get_sorted_df(prod_num_df_7digits,order_mode="sales")
        
        sales_num_df_7digits = sales_num_df.groupby("品種名",as_index=False).sum()
        sales_num_df_7digits = self.get_sorted_df(sales_num_df_7digits,order_mode="sales")
        
        stock_num_df_7digits = stock_num_df.groupby("品種名",as_index=False).sum()
        stock_num_df_7digits = self.get_sorted_df(stock_num_df_7digits,order_mode="sales")
        
        return prod_num_df, sales_num_df, stock_num_df,\
               prod_num_df_7digits, sales_num_df_7digits, stock_num_df_7digits
    
    
    
    def get_prod_hour_df(self,prod_amount_df,cs_df,achieve_rate_df):
        """
        生産時間計算
        
        Args:
            prod_amount_df (_type_): _description_
            cs_df (_type_): _description_
        """
        prod_hour_df = self.convert_amount_to_hour(prod_amount_df,cs_df,achieve_rate_df)
        prod_hour_df_7digits = prod_hour_df.groupby("品種名",as_index=False).sum()
        prod_hour_df_7digits = self.get_sorted_df(prod_hour_df_7digits,order_mode="sales")
        
        
        
        prod_hour_df_factory = prod_hour_df.groupby("工場",as_index=False).sum()
        prod_hour_df_factory = prod_hour_df_factory.sort_values(by='工場', ascending=True)
        
        return prod_hour_df, prod_hour_df_7digits, prod_hour_df_factory
        
    
    def get_ave_width_df(self, prod_amount_df):
        """
        工場毎の平均幅を求める
        各品種の平米を各品種の幅で割る⇒長さが求まる
        
        面積の合計/長さの合計
        
        """
        length_df = prod_amount_df.copy()
        # 辞書の値で各月を割る
        for col in length_df.columns[2:]:  # 月のカラムだけを選択
            length_df[col] = length_df.apply(lambda row: row[col] / self.width_dict.get(row["品種名"], 1), axis=1)
        
        
        length_df_factory = length_df.groupby("工場",as_index=False).sum()
        length_df_factory = length_df_factory.sort_values(by='工場', ascending=True)
        
        prod_amount_df_factory = prod_amount_df.groupby("工場",as_index=False).sum()
        prod_amount_df_factory = prod_amount_df_factory.sort_values(by='工場', ascending=True)
        
        ave_width_df_factory = prod_amount_df_factory.copy()
        ave_width_df_factory[self.jissui_month_list] = prod_amount_df_factory[self.jissui_month_list] /length_df_factory[self.jissui_month_list]
        
        return ave_width_df_factory
        
    
    
    
    
    def get_ave_df(self,df, window_dict):
        """
        〇ヶ月平均のdf
        基本的には販売量のdfがくる
        
        """
        # DataFrameのコピーを作成
        new_df = df.copy()
        
        # 辞書で指定されたウィンドウに基づいて平均を計算
        for month, future_months in window_dict.items():
            new_df[month] = df[future_months].mean(axis=1)

        return new_df
    
    
    def get_monthnum_df(self,dividend_df,divisor_df):
        """
        月数の計算
        dividend_df 割られるdf,
        divisor_df 割るdf

        Args:
            df (_type_): _description_
            df (_type_): _description_
            
        """
        # DataFrameのコピーを作成
        new_df = dividend_df.copy()
        
        new_df[self.jissui_month_list] = dividend_df[self.jissui_month_list] / divisor_df[self.jissui_month_list]
        
        return new_df
        
        
    
    def get_stock_monthnum_df(self,sales_df,stock_df,sales_df_7digits,stock_df_7digits):
        """
        在庫月数

        Args:
            sales_df (_type_): _description_
            stock_df (_type_): _description_
            sales_df_7digits (_type_): _description_
            stock_df_7digits (_type_): _description_

        Returns:
            _type_: _description_
        """
        
        ## 10桁
        ave_sales_df = self.get_ave_df(sales_df,self.ave_month_dict)          #翌〇ヶ月平均販売量
        ave_stock_monthnum_df = self.get_monthnum_df(stock_df,ave_sales_df)   #在庫月数
        
        next_sales_df = self.get_ave_df(sales_df,self.next_month_dict)          #翌月平均販売量
        next_stock_monthnum_df = self.get_monthnum_df(stock_df,next_sales_df)   #在庫月数
        
        ## 7桁
        ave_sales_df_7digits = self.get_ave_df(sales_df_7digits,self.ave_month_dict)          #翌〇ヶ月平均販売量
        ave_stock_monthnum_df_7digits = self.get_monthnum_df(stock_df_7digits,ave_sales_df_7digits)   #在庫月数
        
        next_sales_df_7digits = self.get_ave_df(sales_df_7digits,self.next_month_dict)          #翌月平均販売量
        next_stock_monthnum_df_7digits = self.get_monthnum_df(stock_df_7digits,next_sales_df_7digits)   #在庫月数
        
        
        
        
        return ave_stock_monthnum_df, next_stock_monthnum_df, ave_stock_monthnum_df_7digits, next_stock_monthnum_df_7digits,\
               ave_sales_df, next_sales_df
    
    
    
    
    def get_diff_stock_monthnum_df(self,stock_month_num,basic_stock_df,mode="over"):
        """
        mode=over   超過在庫月数
        mode=shortage 不足在庫月数
        
        
        

        Args:
            stock_month_num (_type_): _description_
            basic_stock_max_df (_type_): _description_

        Returns:
            _type_: _description_
        """
        
        # DataFrameのコピーを作成
        new_df = stock_month_num.copy()
        
        if mode == "over":
            new_df[self.jissui_month_list] = stock_month_num[self.jissui_month_list] - basic_stock_df[self.jissui_month_list]
            return new_df
        if mode == "shortage":
            new_df[self.jissui_month_list] =  basic_stock_df[self.jissui_month_list] - stock_month_num[self.jissui_month_list]
            return new_df
    
    
    def get_diff_prod_amount_df(self,diff_stock_monthnum_df,ave_sales_df,stock_df):
        """
        超過生産量
        平均販売量に超過在庫月数かければそれが超過生産量になる。
        
        ・在庫0で(計算に使う)販売量も0なら、理想の在庫を持てていると考えて、超過生産量は0とする。
        ・在庫30千平米で(計算に使う)販売量が0なら、在庫持ちすぎていると考えて、超過生産量は30千平米とする。


        不足生産量
        平均販売量に不足在庫月数かければそれが不足生産量になる。
        
        ・在庫0で(計算に使う)販売量も0なら、理想の在庫を持てていると考えて、不足生産量は0とする。
        ・在庫-30千平米で(計算に使う)販売量が0なら、在庫不足していると考えて、不足生産量は30千平米とする。





        Args:
            over_stock_monthnum_df (_type_): _description_
            ave_sales_df (_type_): _description_
        """
        # DataFrameのコピーを作成
        new_df = diff_stock_monthnum_df.copy()
        
        new_df[self.jissui_month_list] = ave_sales_df[self.jissui_month_list]*diff_stock_monthnum_df[self.jissui_month_list]
        
        #NaNのばあいは置き換え（在庫-販売量で置き換え）超過の場合も不足の場合も正方向をとるので絶対値する
        new_df[self.jissui_month_list] = new_df[self.jissui_month_list].combine_first(abs(stock_df[self.jissui_month_list] - ave_sales_df[self.jissui_month_list]))

        
        new_df_7digits = new_df.groupby("品種名",as_index=False).sum()
        new_df_7digits = self.get_sorted_df(new_df_7digits,order_mode="sales")
        
        
        return new_df, new_df_7digits
        
    
    
    
    
    def get_cost_df(self,unit_cost_df,prod_amount_df):
        """
        製造限界利益、生産高のデータフレーム作成。
        
        unit_cost_df:FC単価、または（VC単価+FC単価）
        
        
        製造限界利益:実態はFCのこと
          計算式  FC単価（円/㎡）*生産量（千㎡）*1000  =  製造限界利益（円）
        
        生産高:実態はTC（VC+FC）のこと
          計算式  （VC単価+FC単価）*生産量（千㎡）*1000  =  生産高（円）
        
        
        TODO: 10桁レベル、7桁レベル、工場別、でのデータフレームの作成。（ほかの生産量の時とかにどうやってたか参考に）
        
        """
        # DataFrameのコピーを作成
        new_df = unit_cost_df.copy()
        
        #単位：千円
        new_df[self.jissui_month_list] = unit_cost_df[self.jissui_month_list]*prod_amount_df[self.jissui_month_list]
        
        #7桁粒度
        new_df_7digits = new_df.groupby("品種名",as_index=False).sum()
        new_df_7digits = self.get_sorted_df(new_df_7digits,order_mode="sales")
        
        #工場別
        new_df_factory = new_df.groupby("工場",as_index=False).sum()
        new_df_factory = new_df_factory.sort_values(by='工場', ascending=True)
        
        
        return new_df,new_df_7digits,new_df_factory
    
    def get_profit_output_df(self,std_fc_df,std_vc_df,prod_amount_df):
        """
        製造限界利益、生産高のデータフレーム作成。
        
        """
        marginal_profit_df,marginal_profit_df_7digits,\
        marginal_profit_df_factory = self.get_cost_df(std_fc_df,prod_amount_df)
        
        std_tc_df = std_vc_df.copy()
        std_tc_df[self.jissui_month_list] = std_vc_df[self.jissui_month_list] + std_fc_df[self.jissui_month_list]
        
        
        prod_output_df,prod_output_df_7digits,\
        prod_output_df_factory = self.get_cost_df(std_tc_df,prod_amount_df)
        
        return marginal_profit_df,marginal_profit_df_7digits,marginal_profit_df_factory,\
               prod_output_df, prod_output_df_7digits, prod_output_df_factory
        
    
    
    def get_switch_time_df(self,switch_time_dict):
        """
        切り替え時間のデータフレーム作成
        
        """
        nested_dict = {}
        
        # 各キーとその値を取得
        for key in switch_time_dict.keys():
            factory, month = key  # キーをアンパック
            value = switch_time_dict[factory, month].value()  # 各変数の値を取得

            # 入れ子の辞書を作成
            if factory not in nested_dict:
                nested_dict[factory] = {}
            
            nested_dict[factory][month] = value
        
        
        # 入れ子の辞書をデータフレームに変換
        switch_time_df = pd.DataFrame(nested_dict).T
        switch_time_df.reset_index(inplace=True)
        switch_time_df.rename(columns={'index': '工場'}, inplace=True)
        switch_time_df = switch_time_df.sort_values(by='工場', ascending=True)
        
        #st.write(switch_time_df)
        switch_time_df = switch_time_df.reset_index(drop=True)
        
        return switch_time_df
    
    
    # データフレームを作成する関数
    def get_prod_capacity_df(self, key):
        # 工場名と月を基にデータフレームを構築
        factories = list(self.prod_capacity_dict.keys())
        
        # 各工場のデータを抽出し、データフレームを作成
        df_data = {}
        for factory in factories:
            df_data[factory] = [self.prod_capacity_dict[factory].get(month, {}).get(key, None) for month in self.jissui_month_list]
        
        # データフレームに変換
        df = pd.DataFrame(df_data, index=self.jissui_month_list).T
        df.index.name = '工場'
        # インデックス「工場」をカラムに変更
        df_reset = df.reset_index()
        df_reset = df_reset.sort_values(by='工場', ascending=True)
        df_reset = df_reset.reset_index(drop=True)
        
        return df_reset
    
    
    def get_tail_head_df(self,inter_switch_time_tail_df,inter_switch_time_head_df):
        # Create an empty DataFrame to store the results
        df_combined = pd.DataFrame()

        # Iterate over the months (N) and combine N-1 month (from 月末) and N month (from 月初) values
        for i in range(1, len(self.jissui_month_list)):
            # Add the N-1 month value from '月末' and the N month value from '月初'
            df_combined[self.jissui_month_list[i]] = inter_switch_time_tail_df[self.jissui_month_list[i-1]] + inter_switch_time_head_df[self.jissui_month_list[i]]

        # Include the '工場' column from the original data
        df_combined.insert(0, '工場', inter_switch_time_head_df['工場'])
        
        new_column_names = [f"{self.jissui_month_list[i-1]}末 + {self.jissui_month_list[i]}初" for i in range(1, len(self.jissui_month_list))]

        # Assign the new column names to the DataFrame
        df_combined.columns = ['工場'] + new_column_names
        
        df_combined = df_combined.sort_values(by='工場', ascending=True)
        df_combined = df_combined.reset_index(drop=True)
        
        return df_combined
    
    
    
    
    
    
    def get_usage_diff_df(self,prod_hour_df_factory):
        
        #暦日時間のdf
        #保全／工事
        #開発テスト
        #生技テスト
        #停機
        #TODO 立上、立下、切替などなど内訳わかるような出力にする。
        
        
        calendar_day_df = self.get_prod_capacity_df("暦日時間")
        maint_df = self.get_prod_capacity_df("保全")
        dev_test_df = self.get_prod_capacity_df("開発テスト")
        manu_test_df = self.get_prod_capacity_df("生技テスト")
        stopping_df = self.get_prod_capacity_df("計画停止")
        
        
        # fuka_df = pd.DataFrame(self.fuka_dict).T
        # fuka_df.reset_index(inplace=True)
        # fuka_df.rename(columns={'index': '工場'}, inplace=True)
        # fuka_df = fuka_df[["工場"]+self.jissui_month_list]
        # fuka_df = fuka_df.sort_values(by='工場', ascending=True)
        
        # st.write("カレンダー")
        # st.write(calendar_day_df)
        
        #ave_switch_df = pd.DataFrame(self.ave_switch_dict).T
        #ave_switch_df.reset_index(inplace=True)
        #ave_switch_df.rename(columns={'index': '工場'}, inplace=True)
        
        #fuka_diff_switch_df = fuka_df.copy()
        #fuka_diff_switch_df[self.jissui_month_list] = fuka_df[self.jissui_month_list] - ave_switch_df[self.jissui_month_list]
        
        switch_time_df = self.get_switch_time_df(self.switch_time_dict)
        
        inter_switch_time_head_df = self.get_switch_time_df(self.inter_switch_time_head_dict)
        inter_switch_time_tail_df = self.get_switch_time_df(self.inter_switch_time_tail_dict)
        
        inter_switch_time_df = inter_switch_time_head_df.copy()
        inter_switch_time_df[self.jissui_month_list] = inter_switch_time_head_df[self.jissui_month_list] + \
                                                        inter_switch_time_tail_df[self.jissui_month_list]
        
        inter_switch_time_tail_head_df = self.get_tail_head_df(inter_switch_time_tail_df,inter_switch_time_head_df)
        
        
        #負荷時間（暦日時間 - 保全 - 開発テスト - 生技テスト - 計画停止）
        fuka_df = calendar_day_df.copy()
        fuka_df[self.jissui_month_list] = calendar_day_df[self.jissui_month_list] - maint_df[self.jissui_month_list] - \
                                            dev_test_df[self.jissui_month_list] - manu_test_df[self.jissui_month_list] - \
                                            stopping_df[self.jissui_month_list]
        
        #月の総切替時間
        all_switch_time_df = switch_time_df.copy()
        all_switch_time_df[self.jissui_month_list] = switch_time_df[self.jissui_month_list] + inter_switch_time_df[self.jissui_month_list]
        
        
        
        #生産可能時間(暦日時間 - 保全 - 開発テスト - 生技テスト - 計画停止 - 切替-立上-立下)
        fuka_diff_switch_df = fuka_df.copy()
        fuka_diff_switch_df[self.jissui_month_list] = fuka_df[self.jissui_month_list] - switch_time_df[self.jissui_month_list] - inter_switch_time_df[self.jissui_month_list]
        
        
        # 余力時間
        usage_diff_df = fuka_df.copy()
        usage_diff_df[self.jissui_month_list] = fuka_diff_switch_df[self.jissui_month_list] - prod_hour_df_factory[self.jissui_month_list]
        
        
        
        #余力時間をメイン品種の平米換算したもの。（余力時間がマイナスな時重要。）
        
        
        
        
        
        
        
        
        
        
        return fuka_diff_switch_df, usage_diff_df, switch_time_df,inter_switch_time_head_df,inter_switch_time_tail_df,inter_switch_time_df,\
                calendar_day_df, maint_df, dev_test_df, manu_test_df, stopping_df,fuka_df,inter_switch_time_tail_head_df,all_switch_time_df
        

    
    
    
    def get_breakdown_time_record(self,df,plant,name,bundle="normal"):
        """
        {
            "L6":{
            "保全①":{
            "8月":272
            "9月":208
            }
            "保全②":{
            "10月":384
            }
            }
                }
       
        """
        load_time_row = df[df['工場'] == plant].iloc[0]
        
        
        
        if bundle =="diff":
            #抱き合わせ保全の部分は保全分差し引く
            if plant in self.bundle_maint_dict.keys():
                for maint_name in self.bundle_maint_dict[plant].keys():
                    for month in self.bundle_maint_dict[plant][maint_name].keys():
                        load_time_row[month] = load_time_row[month] - self.bundle_maint_dict[plant][maint_name][month]
        
        if bundle =="plus":
            #抱き合わせ保全の部分は保全分足す
            if plant in self.bundle_maint_dict.keys():
                for maint_name in self.bundle_maint_dict[plant].keys():
                    for month in self.bundle_maint_dict[plant][maint_name].keys():
                        load_time_row[month] = load_time_row[month] + self.bundle_maint_dict[plant][maint_name][month]
        
        
        
                
                
                
        load_time_record = {
            "内訳名": name,
            "工場": plant,
            **load_time_row[self.jissui_month_list].to_dict()
        }
        return load_time_record
        
    
    
    
    def get_breakdown_time_df(self,prod_hour_df,fuka_df,all_switch_time_df,usage_diff_df):
        """
        工程時間内訳
        
        """
        
        breakdown_time_df = prod_hour_df.copy()
        breakdown_time_df.rename(columns={"品種名": "内訳名"}, inplace=True)
        
        # 各工場の既存レコードの下に新しいレコードを追加
        updated_df = pd.DataFrame()
        for plant in ["L1","L2","L3","L4","L5","L6","L7"]:
            factory_df = breakdown_time_df[breakdown_time_df['工場'] == plant]
            additional_records = []
            
            record1 = {"内訳名": "暦日時間", "工場": plant}
            record1.update({month: self.prod_capacity_dict[plant][month]["暦日時間"] for month in self.jissui_month_list})
            record1_df = pd.DataFrame([record1])
            
            
            
            
            
            for category in ['保全', '開発テスト', '生技テスト', '計画停止']:
                record = {"内訳名": category, "工場": plant}
                record.update({month: self.prod_capacity_dict[plant][month][category] for month in self.jissui_month_list})
                additional_records.append(record)
            
            
            load_time_record = self.get_breakdown_time_record(all_switch_time_df,plant,"切替時間",bundle="diff")
            additional_records.append(load_time_record)
            
            
            for category in ["立下","立上"]:
                record = {"内訳名": category, "工場": plant}
                record.update({month:0 for month in self.jissui_month_list})
                additional_records.append(record)
                
            
            load_time_record = self.get_breakdown_time_record(fuka_df,plant,"負荷時間")
            additional_records.append(load_time_record)
            
            load_time_record = self.get_breakdown_time_record(usage_diff_df,plant,"余力時間",bundle="plus")
            additional_records.append(load_time_record)
                    
            
            
            
                    
            additional_df = pd.DataFrame(additional_records)
            updated_df = pd.concat([updated_df, record1_df,factory_df, additional_df], ignore_index=True)
            
            #updated_df = pd.concat([updated_df, additional_df], ignore_index=True)
            
        print(updated_df)
        
        
        #['保全', '開発テスト', '生技テスト', '計画停止',"負荷時間","切替時間","立下","立上","余力時間"]:
        
        
        # # 処理開始：各工場ごとに負荷時間を挿入
        # result_df = pd.DataFrame()

        # for plant in self.plant_list:
        #     # 元データから該当する工場のデータを取得
        #     plant_df = updated_df[updated_df['工場'] == plant]
        #     # 既存の負荷時間レコードがない場合のみ追加
        #     if not any(plant_df['内訳名'] == "負荷時間"):
        #         # 負荷時間データを取得
        #         load_time_row = fuka_df[fuka_df['工場'] == plant].iloc[0]
        #         load_time_record = {
        #             "内訳名": "負荷時間",
        #             "工場": plant,
        #             **load_time_row[self.jissui_month_list].to_dict()
        #         }
        #         # 負荷時間を工場データの最後に追加
        #         plant_df = pd.concat([plant_df, pd.DataFrame([load_time_record])], ignore_index=True)
        #     # 結果データフレームに統合
        #     result_df = pd.concat([result_df, plant_df], ignore_index=True)

        # # 結果を表示または保存
        # print(result_df)
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        return updated_df

    
    
    
    
    






    def get_ave_achieve_rate_df(self,achieve_rate_df,prod_hour_df):
        weighted_achieve = achieve_rate_df.copy()
        for month in self.jissui_month_list:
            # Multiply achievement rate by production time
            weighted_achieve[month] = achieve_rate_df[month] *prod_hour_df[month]

        # Group by factory and calculate the weighted sum for each month
        weighted_sum = weighted_achieve.groupby('工場')[self.jissui_month_list].sum()
        total_production_time = prod_hour_df.groupby('工場')[self.jissui_month_list].sum()

        # Calculate average achievement rate by dividing weighted sum by total production time
        ave_achieve_rate_df = weighted_sum / total_production_time
        
        ave_achieve_rate_df.index.name = '工場'
        # インデックス「工場」をカラムに変更
        ave_achieve_rate_df= ave_achieve_rate_df.reset_index()
        ave_achieve_rate_df = ave_achieve_rate_df.sort_values(by='工場', ascending=True)
        ave_achieve_rate_df = ave_achieve_rate_df.reset_index(drop=True)
        
        
        return ave_achieve_rate_df
        

    def get_abnormal_time(self,prod_hour_df_factory,ave_achieve_rate_df):
        """
        異常時間のデータフレーム作成。
        
        """
        
        # DataFrameのコピーを作成
        abnormal_time_df = prod_hour_df_factory.copy()
        
        abnormal_time_df[self.jissui_month_list] = prod_hour_df_factory[self.jissui_month_list]*(1 - ave_achieve_rate_df[self.jissui_month_list])
        
        return abnormal_time_df



    
    def output_excel(self,df, sheet_name, round_num=0,initial_flag=False):
        """
        エクセルファイルとしての書き出し
        一番初めに書き出す場合にinitial_flagをTrueにする
        
        """
        
        df = df.round(round_num)           #
        df = df.replace([np.inf, -np.inf], np.nan)
        
        if initial_flag == True:
            df.to_excel(self.result_file_path,sheet_name=sheet_name,index=None)
        
        else:
            with pd.ExcelWriter(self.result_file_path,engine='openpyxl', mode='a') as writer:
                df.to_excel(writer,sheet_name=sheet_name,index=None)
        
        return 
    
    def output_excel_multi(self, df_dict,sheet_name, round_num=0, space_between=1):
        """
        複数のデータフレームをタイトル付きで同じExcelシートに書き出します。
        
        :param file_name: 出力するExcelファイルの名前
        :param df_dict: データフレームとそのタイトルを持つ辞書
                        キー: タイトル (str)
                        値: データフレーム (pd.DataFrame)
        :param space_between: データフレーム間の空白行数 (デフォルトは2)
        """
        # 既存のワークブックを読み込む
        book = load_workbook(self.result_file_path)
        
        # 新しいシートを追加
        if sheet_name in book.sheetnames:
            raise ValueError(f"シート '{sheet_name}' はすでに存在します。別のシート名を指定してください。")
        worksheet = book.create_sheet(sheet_name)
        
        start_row = 0
        
        # 下線のスタイルを定義
        underline = Border(bottom=Side(style='thin'))
        
        for title, df in df_dict.items():
            # 前処理
            df = df.round(round_num)
            df = df.replace([np.inf, -np.inf], np.nan)
            
            # タイトルを書き込む
            title_cell = worksheet.cell(row=start_row + 1, column=1, value=title)
            title_cell.font = Font(bold=True)
            
            # カラム名を書き込む
            for col_num, col_name in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=start_row + 2, column=col_num, value=col_name)
                cell.font = Font(bold=True)
                cell.border = underline
            
            # データフレームを書き込む
            for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 3):
                for c_idx, value in enumerate(row, start=1):
                    worksheet.cell(row=r_idx, column=c_idx, value=value)
            
            # 次のデータフレームの開始行を計算
            start_row += len(df) + space_between + 3  # データフレームの行数 + タイトル行 + カラム名行 + 空白行数 + 1
        
        # ワークブックを保存
        book.save(self.result_file_path)
        
        return
    
    
    
    
    
    
    
    # データをフラット化する関数
    def flatten_dict(self, d, parent_key='', sep='_'):
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self.flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)
        
    def output_meta_params_info(self,initial_flag=False):
        """
        パラメータのメタ情報をエクセルに書き出す
        """
        
        # フラット化されたデータ
        flat_data = self.flatten_dict(self.meta_params_info_dict)

        # pandas DataFrameに変換
        df = pd.DataFrame([flat_data])
        
        if initial_flag == True:
            df.T.to_excel(self.result_file_path,sheet_name="パラメータメタ情報")
        
        else:
            with pd.ExcelWriter(self.result_file_path,engine='openpyxl', mode='a') as writer:
                    df.T.to_excel(writer,sheet_name="パラメータメタ情報")
        
        
    
    
    
    def output_all(self,cs_df, basic_stock_min_df, basic_stock_max_df, init_stock_df,
                        prod_amount_df, sales_amount_df, stock_amount_df,prod_amount_df_7digits,
                        sales_amount_df_7digits, stock_amount_df_7digits,prod_num_df, sales_num_df, stock_num_df,
                        prod_num_df_7digits, sales_num_df_7digits, stock_num_df_7digits,
                        prod_hour_df, prod_hour_df_7digits, prod_hour_df_factory,
                        ave_stock_monthnum_df, next_stock_monthnum_df,
                        ave_stock_monthnum_df_7digits, next_stock_monthnum_df_7digits,
                        ave_sales_amount_df, next_sales_amount_df,
                        over_stock_monthnum_df,over_prod_amount_df,over_prod_amount_df_7digits,
                        over_prod_hour_df, over_prod_hour_df_7digits,over_prod_hour_df_factory,
                        shortage_stock_monthnum_df,shortage_prod_amount_df,shortage_prod_amount_df_7digits,
                        shortage_prod_hour_df, shortage_prod_hour_df_7digits,shortage_prod_hour_df_factory,
                        marginal_profit_df,marginal_profit_df_7digits,marginal_profit_df_factory,
                        prod_output_df, prod_output_df_7digits, prod_output_df_factory,
                        fuka_diff_switch_df, usage_diff_df,switch_time_df,
                        inter_switch_time_head_df,inter_switch_time_tail_df,inter_switch_time_df,
                        calendar_day_df, maint_df, dev_test_df, manu_test_df, stopping_df,fuka_df,
                        inter_switch_time_tail_head_df,all_switch_time_df,ave_achieve_rate_df,
                        ave_width_df_factory,breakdown_time_df,abnormal_time_df):
        
        self.output_meta_params_info(initial_flag=True)
        
        self.output_excel(breakdown_time_df,"工程時間内訳",round_num=1)
        
        
        df_dict_switch = {
                   "総切替時間":all_switch_time_df,
                   "月中切替時間":switch_time_df,
                   "月間切替時間(前月末＋当月初)":inter_switch_time_tail_head_df,
                   "月間切替時間_月初":inter_switch_time_head_df,
                   "月間切替時間_月末":inter_switch_time_tail_df,
                   "月間切替時間(当月初＋当月末)":inter_switch_time_df}
        self.output_excel_multi(df_dict_switch,"切替時間内訳",round_num=2)
        
        
        
        self.output_excel(ave_achieve_rate_df,"平均生産量達成率",round_num=2)
        self.output_excel(abnormal_time_df,"異常時間",round_num=1)
        self.output_excel(stock_amount_df,"月末在庫",round_num=1)
        #self.output_excel(ave_stock_monthnum_df,f"月末在庫月数({self.ave_month_num}ヶ月平均販売量に対して)",round_num=2)
        self.output_excel(ave_stock_monthnum_df,"月末在庫月数(平均販売量に対して)",round_num=2)
        self.output_excel(over_stock_monthnum_df,"超過月末在庫月数(平均販売量に対して)",round_num=2)
        
        self.output_excel(basic_stock_min_df,"基準在庫月数Min",round_num=2)
        self.output_excel(basic_stock_max_df,"基準在庫月数Max",round_num=2)
        
        
        self.output_excel(ave_sales_amount_df,"平均販売量",round_num=1)
        self.output_excel(prod_amount_df,"生産量",round_num=1)
        self.output_excel(prod_hour_df,"生産時間",round_num=1)
        self.output_excel(sales_amount_df,"販売量",round_num=1)
        self.output_excel(init_stock_df,"期首在庫")
        
        self.output_excel(stock_amount_df_7digits,"7桁月末在庫",round_num=1)
        self.output_excel(ave_stock_monthnum_df_7digits,"7桁月末在庫月数(平均販売量に対して)",round_num=1)
        self.output_excel(prod_amount_df_7digits,"7桁生産量",round_num=1)
        self.output_excel(sales_amount_df_7digits,"7桁販売量",round_num=1)
        self.output_excel(prod_hour_df_7digits,"7桁生産時間",round_num=1)
        #self.output_excel(prod_hour_df_factory,"工場毎生産時間",round_num=1)
        #self.output_excel(fuka_diff_switch_df,"生産可能時間",round_num=1)
        #self.output_excel(usage_diff_df,"余力時間",round_num=1)
        
        df_dict = {"暦日時間":calendar_day_df,
                   "負荷時間（暦日時間 - 保全 - 開発テスト - 生技テスト - 計画停止）":fuka_df,
                   "総切替時間":all_switch_time_df,
                   "月間切替時間(前月末＋当月初)":inter_switch_time_tail_head_df,
                   "月間切替時間(当月初＋当月末)":inter_switch_time_df,
                   "月中切替時間":switch_time_df,
                   "生産可能時間(負荷時間 - 月中切替時間 - 月間切替時間)":fuka_diff_switch_df,
                   "使用生産時間":prod_hour_df_factory,
                   "余力時間(生産可能時間 - 使用生産時間)":usage_diff_df,
                   "月間切替時間_月初":inter_switch_time_head_df,
                   "月間切替時間_月末":inter_switch_time_tail_df}
        self.output_excel_multi(df_dict,"工場毎生産時間",round_num=2)
        
        self.output_excel(ave_width_df_factory,"工場毎平均幅",round_num=3)
        
        self.output_excel(cs_df,"cs")

        
        self.output_excel(next_stock_monthnum_df_7digits,"7桁月末在庫月数(翌月販売量に対して)",round_num=1)
        self.output_excel(next_stock_monthnum_df,"月末在庫月数(翌月販売量に対して)",round_num=1)
        

        
        self.output_excel(prod_num_df,"生産本数(3900m換算)")
        self.output_excel(sales_num_df,"販売本数(3900m換算)")
        self.output_excel(stock_num_df,"月末在庫本数(3900m換算)")
        
        
        self.output_excel(prod_num_df_7digits,"7桁生産本数(3900m換算)")
        self.output_excel(sales_num_df_7digits,"7桁販売本数(3900m換算)")
        self.output_excel(stock_num_df_7digits,"7桁月末在庫本数(3900m換算)")
        
        
        
        
        #TODO 必要性が高まったところで順番考える
        #self.output_excel(over_stock_monthnum_df,f"超過月末在庫月数(平均販売量に対して)",round_num=1)
        self.output_excel(over_prod_amount_df,"超過生産量(平均販売量に対して)")
        self.output_excel(over_prod_amount_df_7digits,"7桁超過生産量(平均販売量に対して)")
        self.output_excel(over_prod_hour_df,"超過生産時間(平均販売量に対して)")
        self.output_excel(over_prod_hour_df_7digits,"7桁超過生産時間(平均販売量に対して)")
        self.output_excel(over_prod_hour_df_factory,"工場毎超過生産時間(平均販売量に対して)")
        
        self.output_excel(shortage_stock_monthnum_df,"不足月末在庫月数(平均販売量に対して)",round_num=1)
        self.output_excel(shortage_prod_amount_df,"不足生産量(平均販売量に対して)")
        self.output_excel(shortage_prod_amount_df_7digits,"7桁不足生産量(平均販売量に対して)")
        self.output_excel(shortage_prod_hour_df,"不足生産時間(平均販売量に対して)")
        self.output_excel(shortage_prod_hour_df_7digits,"7桁不足生産時間(平均販売量に対して)")
        self.output_excel(shortage_prod_hour_df_factory,"工場毎不足生産時間(平均販売量に対して)")
        
 
        self.output_excel(marginal_profit_df, "製造限界利益",round_num=1)
        self.output_excel(marginal_profit_df_7digits, "7桁製造限界利益",round_num=1)
        self.output_excel(marginal_profit_df_factory, "工場毎製造限界利益",round_num=1)
        
        self.output_excel(prod_output_df, "生産高",round_num=1)
        self.output_excel(prod_output_df_7digits, "7桁生産高",round_num=1)
        self.output_excel(prod_output_df_factory, "工場毎生産高",round_num=1)
        
    
    
    
    def main(self):
        #パラメータ類もdf化
        cs_df, basic_stock_min_df, basic_stock_max_df, init_stock_df,std_fc_df,std_vc_df,achieve_rate_df = self.get_params_df()
        
        #basic_stock_max_monthnum_df = self.convert_daynum_to_monthnum(basic_stock_max_df)    #基準在庫月数Max
        #basic_stock_min_monthnum_df = self.convert_daynum_to_monthnum(basic_stock_min_df)    #基準在庫月数Min
        
        #10桁・7桁生産量販売量在庫（千平米）
        prod_amount_df, sales_amount_df, stock_amount_df,\
        prod_amount_df_7digits, sales_amount_df_7digits, stock_amount_df_7digits = self.get_amount_df(init_stock_df)
        
        #10桁・7桁生産量販売量在庫（本数）
        prod_num_df, sales_num_df, stock_num_df,\
        prod_num_df_7digits, sales_num_df_7digits, stock_num_df_7digits = self.get_num_df(prod_amount_df,
                                                                                          sales_amount_df,
                                                                                          stock_amount_df)
        
        
        #平均幅(工場毎月毎)
        ave_width_df_factory = self.get_ave_width_df(prod_amount_df)
        
        
        
        
        
        #生産時間
        prod_hour_df, prod_hour_df_7digits, prod_hour_df_factory = self.get_prod_hour_df(prod_amount_df,cs_df,achieve_rate_df)
        
        
        #稼働率
        fuka_diff_switch_df, usage_diff_df,switch_time_df,\
        inter_switch_time_head_df,inter_switch_time_tail_df,inter_switch_time_df,\
        calendar_day_df, maint_df, dev_test_df, manu_test_df, stopping_df,fuka_df,\
        inter_switch_time_tail_head_df,all_switch_time_df= self.get_usage_diff_df(prod_hour_df_factory)   #負荷時間-切り替え時間、余力時間
        
        
        
        #工程時間内訳（工場毎月毎）
        breakdown_time_df = self.get_breakdown_time_df(prod_hour_df,fuka_df,all_switch_time_df,usage_diff_df)
        
        
        
        #翌月のみ・翌月〇ヶ月平均販売量を基にした在庫月数
        ave_stock_monthnum_df, next_stock_monthnum_df, \
        ave_stock_monthnum_df_7digits, next_stock_monthnum_df_7digits,\
        ave_sales_amount_df, next_sales_amount_df                     = self.get_stock_monthnum_df(sales_amount_df,
                                                                                                   stock_amount_df,
                                                                                                   sales_amount_df_7digits,
                                                                                                   stock_amount_df_7digits)
        
        
        #超過在庫月数
        over_stock_monthnum_df = self.get_diff_stock_monthnum_df(ave_stock_monthnum_df,basic_stock_max_df,mode="over") 
        
        
        
        #超過生産量
        over_prod_amount_df,over_prod_amount_df_7digits = self.get_diff_prod_amount_df(over_stock_monthnum_df,ave_sales_amount_df,stock_amount_df)
        
        #超過生産時間
        over_prod_hour_df, over_prod_hour_df_7digits, \
        over_prod_hour_df_factory = self.get_prod_hour_df(over_prod_amount_df,cs_df,achieve_rate_df) 
        
        
        
        #不足在庫月数
        shortage_stock_monthnum_df = self.get_diff_stock_monthnum_df(ave_stock_monthnum_df,basic_stock_min_df,
                                                                     mode="shortage")
        
        #不足生産量
        shortage_prod_amount_df,shortage_prod_amount_df_7digits = self.get_diff_prod_amount_df(shortage_stock_monthnum_df,
                                                                                       ave_sales_amount_df,
                                                                                       stock_amount_df)
        
        #不足生産時間
        shortage_prod_hour_df, shortage_prod_hour_df_7digits, \
        shortage_prod_hour_df_factory = self.get_prod_hour_df(shortage_prod_amount_df,cs_df,achieve_rate_df) 
        
        #製造限界利益、生産高
        marginal_profit_df,marginal_profit_df_7digits,marginal_profit_df_factory,\
        prod_output_df, prod_output_df_7digits, prod_output_df_factory = self.get_profit_output_df(std_fc_df,std_vc_df,prod_amount_df)
 
        
        #平均生産量達成率
        ave_achieve_rate_df = self.get_ave_achieve_rate_df(achieve_rate_df,prod_hour_df)
        
        
        #異常時間
        abnormal_time_df = self.get_abnormal_time(prod_hour_df_factory,ave_achieve_rate_df)
        
        

        #ファイル書き出し
        self.output_all(cs_df, basic_stock_min_df, basic_stock_max_df, init_stock_df,
                        prod_amount_df, sales_amount_df, stock_amount_df,prod_amount_df_7digits,
                        sales_amount_df_7digits, stock_amount_df_7digits,prod_num_df, sales_num_df, stock_num_df,
                        prod_num_df_7digits, sales_num_df_7digits, stock_num_df_7digits,
                        prod_hour_df, prod_hour_df_7digits, prod_hour_df_factory,
                        ave_stock_monthnum_df, next_stock_monthnum_df,
                        ave_stock_monthnum_df_7digits, next_stock_monthnum_df_7digits,
                        ave_sales_amount_df, next_sales_amount_df,
                        over_stock_monthnum_df,over_prod_amount_df,over_prod_amount_df_7digits,
                        over_prod_hour_df, over_prod_hour_df_7digits,over_prod_hour_df_factory,
                        shortage_stock_monthnum_df,shortage_prod_amount_df,shortage_prod_amount_df_7digits,
                        shortage_prod_hour_df, shortage_prod_hour_df_7digits,shortage_prod_hour_df_factory,
                        marginal_profit_df,marginal_profit_df_7digits,marginal_profit_df_factory,
                        prod_output_df, prod_output_df_7digits, prod_output_df_factory,
                        fuka_diff_switch_df, usage_diff_df,switch_time_df,inter_switch_time_head_df,
                        inter_switch_time_tail_df,inter_switch_time_df,
                        calendar_day_df, maint_df, dev_test_df, manu_test_df, stopping_df,fuka_df,
                        inter_switch_time_tail_head_df,all_switch_time_df,ave_achieve_rate_df,ave_width_df_factory,
                        breakdown_time_df,abnormal_time_df
                        )
        
        
        
    
        # with open("./test/sales_amount_df.pickle", mode='wb') as f:
        #         pickle.dump(sales_amount_df,f)
        # with open("./test/init_stock_df.pickle", mode='wb') as f:
        #         pickle.dump(init_stock_df,f)
        
        return init_stock_df, sales_amount_df
        
####################################################################################################################
if __name__ == "__main__":
    with open('./test/all_params_dict.pickle', 'rb') as f:
        all_params_dict = pickle.load(f)
    
    with open("./結果/mip_result_dict.pickle", 'rb') as f:
        variables = pickle.load(f)
    
    problem = None
    
    filename = "あああ.xlsx"
    
    ResultOutputter(all_params_dict,variables, problem,
                 filename).main()