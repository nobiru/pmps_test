import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from openpyxl.styles import Border, Side, Font, Color
from openpyxl.styles import PatternFill

from output_config import RESULT_DIR


class XlNavigationMaker():
    def __init__(self,result_file_name):
        self.file_path = f"{str(RESULT_DIR)}/{result_file_name}"
        
        self.file_path = RESULT_DIR / result_file_name
        
        
        
        #リンク文字列とシート名の対応
        self.sheet_dict = {"月末在庫":"月末在庫",
                            "月末在庫月数":"月末在庫月数(平均販売量に対して)",
                            "超過月末在庫月数":"超過月末在庫月数(平均販売量に対して)",
                            "基準在庫月数Min":"基準在庫月数Min",
                            "基準在庫月数Max":"基準在庫月数Max",
                            "平均販売量":"平均販売量",
                            "生産量":"生産量",
                            "生産時間":"生産時間",
                            "販売量":"販売量",
                            "期首在庫":"期首在庫",
                            "7桁月末在庫":"7桁月末在庫",
                            "7桁月末在庫月数":"7桁月末在庫月数(平均販売量に対して)",
                            "7桁生産量":"7桁生産量",
                            "7桁販売量":"7桁販売量",
                            "7桁生産時間":"7桁生産時間",
                            "工場毎生産時間":"工場毎生産時間",
                            "cs":"cs",
                            "超過生産量":"超過生産量(平均販売量に対して)",
                            "7桁超過生産量":"7桁超過生産量(平均販売量に対して)",
                            "超過生産時間":"超過生産時間(平均販売量に対して)",
                            "7桁超過生産時間":"7桁超過生産時間(平均販売量に対して)",
                            "工場毎超過生産時間":"工場毎超過生産時間(平均販売量に対して)",
                            "不足月末在庫月数":"不足月末在庫月数(平均販売量に対して)",
                            "不足生産量":"不足生産量(平均販売量に対して)",
                            "7桁不足生産量":"7桁不足生産量(平均販売量に対して)",
                            "不足生産時間":"不足生産時間(平均販売量に対して)",
                            "7桁不足生産時間":"7桁不足生産時間(平均販売量に対して)",
                            "工場毎不足生産時間":"工場毎不足生産時間(平均販売量に対して)"}
        
        self.sheet_dict_swap = {v: k for k, v in self.sheet_dict.items()}
        
        
        #TODOこれ、全体の場所を丸ごと変えたいときに終わるのでいい感じにしたい。（何かからの相対位置にするなど）
        self.link_positions = {"月末在庫":(13,17),
                            "月末在庫月数":(14,17),
                            "超過月末在庫月数":(15,17),
                            "基準在庫月数Min":(18,17),
                            "基準在庫月数Max":(19,17),
                            "平均販売量":(21,17),
                            "生産量":(6,17),
                            "生産時間":(7,17),
                            "販売量":(20,17),
                            "期首在庫":(17,17),
                            "7桁月末在庫":(13,18),
                            "7桁月末在庫月数":(14,18),
                            "7桁生産量":(6,18),
                            "7桁販売量":(20,18),
                            "7桁生産時間":(7,18),
                            "工場毎生産時間":(7,19),
                            "cs":(12,17),
                            "超過生産量":(8,17),
                            "7桁超過生産量":(8,18),
                            "超過生産時間":(9,17),
                            "7桁超過生産時間":(9,18),
                            "工場毎超過生産時間":(9,19),
                            "不足月末在庫月数":(16,17),
                            "不足生産量":(10,17),
                            "7桁不足生産量":(10,18),
                            "不足生産時間":(11,17),
                            "7桁不足生産時間":(11,18),
                            "工場毎不足生産時間":(11,19)}
        
        self.sheet_name_list = list(self.sheet_dict.values())
        
        
        self.params_list = ["cs","期首在庫","基準在庫月数Min",
                            "基準在庫月数Max","7桁販売量"]
        
    
    
    # 列の幅を調整するための関数
    def adjust_column_width(self,sheet, col, text):
        width = max(8, len(text) * 1.2)  # 基本的な幅を計算（最小幅を8とする）
        width=18
        sheet.column_dimensions[get_column_letter(col)].width = width
    
    
    
    def make_format(self,ws):
        ws["Q5"] = "10桁"
        ws["R5"] = "7桁"
        ws["S5"] = "工場毎"
        ws["P6"] = "生産"
        ws["P13"] = "在庫"
        ws["P20"] = "販売"
        return ws
        
        
    def write_border(self,ws, start, end, target, mode="horizontal"):
        # 下罫線のスタイルを定義
        bottom_border = Border(bottom=Side(style='thin'))
        
        
        if mode=="horizontal":
            for col in range(start, end + 1):
                cell = ws.cell(row=target, column=col)
                cell.border = bottom_border
        
        return ws
    

    
    def main(self):
        
        # ワークブックを読み込む
        wb = load_workbook(self.file_path)
        
        
        
        
        # 下罫線を追加する行と列の範囲を設定
        target_row = 5  # 5行目
        start_column = 17  # 17列目
        end_column = 20  # 20列目
        
        
        
        # 各シートにリンクを追加し、列幅を調整
        for sheet_name in self.sheet_name_list+["パラメータメタ情報"]:
            ws = wb[sheet_name]
            for target_sheet in self.sheet_name_list:
                link_row, link_col = self.link_positions[self.sheet_dict_swap[target_sheet]]
                cell_text = self.sheet_dict_swap[target_sheet]
                cell = ws.cell(row=link_row, column=link_col)
                # リンクの設定先を対象シートの指定されたセル位置に設定
                #()が特殊文字扱いされるので、全体をシングルクォーテーションで囲む
                cell.hyperlink = f"#'{target_sheet}'!{get_column_letter(link_col)}{link_row}"
                cell.value = cell_text
                
                # フォントの色を青に設定
                blue_font = Font(color="0000FF")
                cell.font = blue_font
                
                if cell_text in self.params_list:
                    green_fill = PatternFill(start_color='EBF1DE',
                                end_color='EBF1DE',
                                fill_type='solid')
                    cell.fill = green_fill
                
                ws = self.make_format(ws)
                
                # 列幅をテキストの長さに合わせて調整
                self.adjust_column_width(ws, link_col, cell_text)
                
                ws = self.write_border(ws,16,19,5)
                ws = self.write_border(ws,16,19,12)
                ws = self.write_border(ws,16,19,19)
                ws = self.write_border(ws,16,19,21)
            
            
            #品種名のカラム広げる
            column_number = None
            for col in ws.iter_cols(1, ws.max_column):
                for cell in col:
                    if cell.value == '品種名':
                        column_number = cell.column  # openpyxlは1から数え始める
                        
                        # 列幅をテキストの長さに合わせて調整
                        column_letter = openpyxl.utils.get_column_letter(column_number)
                        ws.column_dimensions[column_letter].width = 23
                    
                    if cell.value == '工場':
                        column_number = cell.column  # openpyxlは1から数え始める
                        
                        # 列幅をテキストの長さに合わせて調整
                        column_letter = openpyxl.utils.get_column_letter(column_number)
                        ws.column_dimensions[column_letter].width = 6
            
        
        # 変更を上書き保存
        wb.save(self.file_path)